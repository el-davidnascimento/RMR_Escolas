from datetime import datetime
import openpyxl
import pandas as pd
import datetime as date

############################################# 1 - Reclamacoes ##########################################################

caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\1 - Reclamacoes.xlsx'
nome_aba = '1 - Reclamacoes'
dados_excel = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

# Obter a data atual
data_atual = date.datetime.today()

# Criar uma nova coluna com a data atual
dados_excel['Date_insert'] = data_atual
if pd.api.types.is_datetime64_any_dtype(dados_excel['Data']):
    dados_excel['Data de abertura'] = pd.to_datetime(dados_excel['Data'])
    dados_excel['Mês_abertura'] = dados_excel['Data'].dt.month.map("{:02d}".format)
    dados_excel['Ano_abertura'] = dados_excel['Data'].dt.year
else:
    dados_excel['Data de abertura'] = None
    dados_excel['Mês_abertura'] = None
    dados_excel['Ano_abertura'] = None


dados_excel.loc[dados_excel['Turno'] != 'Integral', 'Turno'] = 'Regular'

Tratado_reclamacoes = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\13.2.2.2.2.1.2.1.2. Tratado\Tratado_reclamacao.xlsx'
dados_excel.to_excel(Tratado_reclamacoes,index=False)

############################################# Tratamento_Reclamacoes
############################# INDICADOR : ocorrências dentro do SLA/ ocorrências total

caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\13.2.2.2.2.1.2.1.2. Tratado\Tratado_reclamacao.xlsx'
dados_excel = pd.read_excel(caminho_arquivo)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_retencao = Aba_metricas['A6'].value
Valor_area_responsavel_retencao = Aba_metricas['B6'].value
Valor_setor_responsavel_retencao = Aba_metricas['C6'].value
Valor_frente2_retencao = Aba_metricas['D6'].value
Valor_indicador_retencao = Aba_metricas['E6'].value
Valor_meta_retencao = Aba_metricas['F6'].value
Valor_ruim_retencao = Aba_metricas['G6'].value
Valor_regular_retencao = Aba_metricas['H6'].value
Valor_otimo_retencao = Aba_metricas['I6'].value
Valor_peso_retencao = Aba_metricas['J6'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_retencao,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_retencao,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_retencao,
    'FRENTE 2': Valor_frente2_retencao,
    'INDICADOR': Valor_indicador_retencao,
    'META': Valor_meta_retencao,
    'RUIM': Valor_ruim_retencao,
    'REGULAR': Valor_regular_retencao,
    'ÓTIMO': Valor_otimo_retencao,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_retencao,
}

# Criar o DataFrame
df = pd.DataFrame(dados)

# Inicializar as listas
Qtde_Reclamacao = []
Qtde_Insatisfacao = []
Qtde_Valor_total = []
Qtde_Valor = []


for mes in meses:
    if dados_excel['Mês_abertura'].isna().all():
        qtde_reclamacao = None
        qtde_insatisfacao = None
    else:
        qtde_reclamacao = dados_excel[
            (dados_excel['Ocorrências (Reclamação ou Insatisfação)'] == 'Reclamação') & (dados_excel['Mês_abertura'] == mes)][
            'Numero de matricula'].count()
        qtde_insatisfacao = dados_excel[
            (dados_excel['Ocorrências (Reclamação ou Insatisfação)'] == 'Insatisfação') & (dados_excel['Mês_abertura'] == mes)][
            'Numero de matricula'].count()

    Qtde_Reclamacao.append(qtde_reclamacao)
    Qtde_Insatisfacao.append(qtde_insatisfacao)


Qtde_Valor = Qtde_Reclamacao


# Atribuir as listas ao DataFrame
df['Valor'] = Qtde_Valor

Tratamento_Reclamacao = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Reclamacao.xlsx'
df.to_excel(Tratamento_Reclamacao,index=False)

############################################# 2 - Atrasos_Faltas_Colaboradores ##########################################################
############################################# Tabela_Faltas_Colaboradores
############################# INDICADOR : dias faltosos de todos os colaboradores/diasde trabalho de todos os colaboradores


caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\2 - Atraso_Faltas_Colaboradores.xlsx'
nome_aba = '2 - Atraso'
dados_excel = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_falta_colaborador = Aba_metricas['A8'].value
Valor_area_responsavel_falta_colaborador = Aba_metricas['B8'].value
Valor_setor_responsavel_falta_colaborador = Aba_metricas['C8'].value
Valor_frente2_falta_colaborador = Aba_metricas['D8'].value
Valor_indicador_falta_colaborador = Aba_metricas['E8'].value
Valor_meta_falta_colaborador = Aba_metricas['F8'].value
Valor_ruim_falta_colaborador = Aba_metricas['G8'].value
Valor_regular_falta_colaborador = Aba_metricas['H8'].value
Valor_otimo_falta_colaborador = Aba_metricas['I8'].value
Valor_peso_falta_colaborador = Aba_metricas['J8'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_falta_colaborador,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_falta_colaborador,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_falta_colaborador,
    'FRENTE 2': Valor_frente2_falta_colaborador,
    'INDICADOR': Valor_indicador_falta_colaborador,
    'META': Valor_meta_falta_colaborador,
    'RUIM': Valor_ruim_falta_colaborador,
    'REGULAR': Valor_regular_falta_colaborador,
    'ÓTIMO': Valor_otimo_falta_colaborador,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_falta_colaborador,
}

# Criar o DataFrame
df = pd.DataFrame(dados)

# Inicializar as listas
Qtde_Valor = []

for mes in meses:
    if dados_excel['Mês'].isna().all():
        Qtde_Valor = None
    else:
        qtde_falta_total = dados_excel.loc[dados_excel['Mês'] == mes, 'Dias de falta'].sum()
        qtde_dias_trabalho_total = dados_excel.loc[dados_excel['Mês'] == mes, 'Dias úteis totais de trabalho'].sum()

        if qtde_dias_trabalho_total != 0:
            Qtde_Valor.append(qtde_falta_total / qtde_dias_trabalho_total)
        else:
            Qtde_Valor.append(None)

Qtde_Valor_Serie = pd.Series(Qtde_Valor)

Qtde_Valor_formatado = Qtde_Valor_Serie.map('{:.2%}'.format)

# Atribuir as listas ao DataFrame
df['Valor'] = Qtde_Valor_formatado

Tratamento_Falta = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Falta.xlsx'
df.to_excel(Tratamento_Falta,index=False)

############################################# Tabela_Atraso_Colaboradores
############################# INDICADOR : dias de atraso de todos os colaboradores/diasde trabalho de todos os colaboradores

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_atraso_colaborador = Aba_metricas['A7'].value
Valor_area_responsavel_atraso_colaborador = Aba_metricas['B7'].value
Valor_setor_responsavel_atraso_colaborador = Aba_metricas['C7'].value
Valor_frente2_atraso_colaborador = Aba_metricas['D7'].value
Valor_indicador_atraso_colaborador = Aba_metricas['E7'].value
Valor_meta_atraso_colaborador = Aba_metricas['F7'].value
Valor_ruim_atraso_colaborador = Aba_metricas['G7'].value
Valor_regular_atraso_colaborador = Aba_metricas['H7'].value
Valor_otimo_atraso_colaborador = Aba_metricas['I7'].value
Valor_peso_atraso_colaborador = Aba_metricas['J7'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_atraso_colaborador,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_atraso_colaborador,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_atraso_colaborador,
    'FRENTE 2': Valor_frente2_atraso_colaborador,
    'INDICADOR': Valor_indicador_atraso_colaborador,
    'META': Valor_meta_atraso_colaborador,
    'RUIM': Valor_ruim_atraso_colaborador,
    'REGULAR': Valor_regular_atraso_colaborador,
    'ÓTIMO': Valor_otimo_atraso_colaborador,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_atraso_colaborador,
}

# Criar o DataFrame
df = pd.DataFrame(dados)

# Inicializar as listas
Qtde_Valor = []

for mes in meses:

    if dados_excel['Mês'].isna().all():
        Qtde_Valor = None
    else:
        qtde_atraso_total = dados_excel.loc[dados_excel['Mês'] == mes, 'Dias úteis com atraso'].sum()
        qtde_dias_trabalho_total = dados_excel.loc[dados_excel['Mês'] == mes, 'Dias úteis totais de trabalho'].sum()

        if qtde_dias_trabalho_total != 0:
            Qtde_Valor.append(qtde_atraso_total / qtde_dias_trabalho_total)
        else:
            Qtde_Valor.append(None)


Qtde_Valor_Serie = pd.Series(Qtde_Valor)

Qtde_Valor_formatado = Qtde_Valor_Serie.map('{:.2%}'.format)

# Atribuir as listas ao DataFrame
df['Valor'] = Qtde_Valor_formatado


Tratamento_Atraso = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Atraso.xlsx'
df.to_excel(Tratamento_Atraso,index=False)

############################################# 3 - Desistencia_de_alunos ##########################################################

caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\3 - Desistencia_de_alunos.xlsx'
nome_aba = '3 - Desistencia de alunos'
dados_excel = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

# Obter a data atual
data_atual = date.datetime.today()

# Criar uma nova coluna com a data atual
dados_excel['Date_insert'] = data_atual
dados_excel.loc[dados_excel['Turno'] != 'Integral', 'Turno'] = 'Regular'
dados_excel['Data de desistência'] = pd.to_datetime(dados_excel['Data de desistência'])
dados_excel['Ano_desistencia'] = dados_excel['Data de desistência'].dt.year
dados_excel['Mês'] = dados_excel['Data de desistência'].dt.month


########################################### ALUNOS TOTAL ################################

caminho_arquivo_alunos_total = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\11 - Listagem_dos_alunos.xlsx'
nome_aba_alunos_total = '15 - Listagem dos alunos - atua'
dados_excel_alunos_total = pd.read_excel(caminho_arquivo_alunos_total, sheet_name=nome_aba_alunos_total)

############################################# Tabela desistencia Regular
############################# INDICADOR : nº matrículas turno regular que desistiram/ nº alunos matriculados
# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_desistencia_regular = Aba_metricas['A4'].value
Valor_area_responsavel_desistencia_regular = Aba_metricas['B4'].value
Valor_setor_responsavel_desistencia_regular = Aba_metricas['C4'].value
Valor_frente2_desistencia_regular = Aba_metricas['D4'].value
Valor_indicador_desistencia_regular = Aba_metricas['E4'].value
Valor_meta_desistencia_regular = Aba_metricas['F4'].value
Valor_ruim_desistencia_regular = Aba_metricas['G4'].value
Valor_regular_desistencia_regular = Aba_metricas['H4'].value
Valor_otimo_desistencia_regular = Aba_metricas['I4'].value
Valor_peso_desistencia_regular = Aba_metricas['J4'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_desistencia_regular,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_desistencia_regular,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_desistencia_regular,
    'FRENTE 2': Valor_frente2_desistencia_regular,
    'INDICADOR': Valor_indicador_desistencia_regular,
    'META': Valor_meta_desistencia_regular,
    'RUIM': Valor_ruim_desistencia_regular,
    'REGULAR': Valor_regular_desistencia_regular,
    'ÓTIMO': Valor_otimo_desistencia_regular,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_desistencia_regular,
}

# Criar o DataFrame
df = pd.DataFrame(dados)

# Inicializar as listas
Desistencia_Regular = []

# Calcular o total de alunos na disciplina
Qtde_total_alunos_disciplina = dados_excel_alunos_total['Numero de Matrícula'].count()

# Calcular o total de alunos na disciplina
Qtde_total_alunos_disciplina = dados_excel_alunos_total['Numero de Matrícula'].count()

for mes in meses:
    if dados_excel['Mês'].isna().all() or Qtde_total_alunos_disciplina == 0:
        Desistencia_Regular = None
    else:
        # Filtrar os dados para o mês atual e o ano atual
        dados_mes_atual = dados_excel[(dados_excel['Mês'] == mes) & (dados_excel['Ano_desistencia'] == ano_atual)]

        # Calcular a quantidade de desistentes do turno Regular no mês atual
        qtde_Desistente_Regular_total = dados_mes_atual[dados_mes_atual['Turno'] == 'Regular'].shape[0]
        medidas_todo_mes_regular = dados_mes_atual[dados_mes_atual['Segmento'] == 'TODOS'].count()

        # Tem o campo de "não houve desistente" informado
        if (medidas_todo_mes_regular == 1).any():
            # Calcular a taxa de desistência regular e adicionar à lista
            if Qtde_total_alunos_disciplina != 0:
                taxa_desistencia = qtde_Desistente_Regular_total / Qtde_total_alunos_disciplina
            else:
                taxa_desistencia = 0
            Desistencia_Regular.append(taxa_desistencia)
        else:
            Desistencia_Regular.append(0)

# Atribuir as listas ao DataFrame
df['Valor'] = Desistencia_Regular

Tratamento_Desistencia_Regular = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Desistencia_Regular.xlsx'
df.to_excel(Tratamento_Desistencia_Regular,index=False)

############################################# Tabela desistencia Integral
############################# INDICADOR : nº matrículas turno integral que desistiram (sem ser por bloqueamento por inadimplência) / nº alunos matriculados
# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_desistencia_integral = Aba_metricas['A5'].value
Valor_area_responsavel_desistencia_integral = Aba_metricas['B5'].value
Valor_setor_responsavel_desistencia_integral = Aba_metricas['C5'].value
Valor_frente2_desistencia_integral = Aba_metricas['D5'].value
Valor_indicador_desistencia_integral = Aba_metricas['E5'].value
Valor_meta_desistencia_integral = Aba_metricas['F5'].value
Valor_ruim_desistencia_integral = Aba_metricas['G5'].value
Valor_regular_desistencia_integral = Aba_metricas['H5'].value
Valor_otimo_desistencia_integral = Aba_metricas['I5'].value
Valor_peso_desistencia_integral = Aba_metricas['J5'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_desistencia_integral,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_desistencia_integral,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_desistencia_integral,
    'FRENTE 2': Valor_frente2_desistencia_integral,
    'INDICADOR': Valor_indicador_desistencia_integral,
    'META': Valor_meta_desistencia_integral,
    'RUIM': Valor_ruim_desistencia_integral,
    'REGULAR': Valor_regular_desistencia_integral,
    'ÓTIMO': Valor_otimo_desistencia_integral,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_desistencia_integral,
}

# Criar o DataFrame
df = pd.DataFrame(dados)

# Inicializar as listas
Desistencia_Integral = []

# Calcular o total de alunos na disciplina
Qtde_total_alunos_disciplina = dados_excel_alunos_total['Numero de Matrícula'].count()

for mes in meses:
    if dados_excel['Mês'].isna().all() or Qtde_total_alunos_disciplina == 0:
        Desistencia_Integral = None
    else:
        # Filtrar os dados para o mês atual e o ano atual
        dados_mes_atual = dados_excel[(dados_excel['Mês'] == mes) & (dados_excel['Ano_desistencia'] == ano_atual)]

        # Calcular a quantidade de desistentes do turno Regular no mês atual
        qtde_Desistente_Integral_total = dados_mes_atual[dados_mes_atual['Turno'] == 'Integral'].shape[0]
        medidas_todo_mes_integral = dados_mes_atual[dados_mes_atual['Segmento'] == 'TODOS'].count()

        # Tem o campo de "não houve desistente" informado
        if  (medidas_todo_mes_integral == 1).any():
            # Calcular a taxa de desistência regular e adicionar à lista
            if Qtde_total_alunos_disciplina != 0:
                taxa_desistencia_integral = qtde_Desistente_Integral_total / Qtde_total_alunos_disciplina
            else:
                taxa_desistencia = 0
            Desistencia_Integral.append(taxa_desistencia_integral)
        else:
            Desistencia_Integral.append(0)

df['Valor'] = Desistencia_Integral

Tratamento_Desistencia_Integral = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Desistencia_Integral.xlsx'
df.to_excel(Tratamento_Desistencia_Integral,index=False)

############################################# 5 - Contratos ##########################################################
############################# INDICADOR : contratos assinados/total de contratos

caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\5 - Contratos.xlsx'
nome_aba = '5 - Contratos'
dados_excel = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

# Obter a data atual
data_atual = date.datetime.today()

# Criar uma nova coluna com a data atual
dados_excel['Date_insert'] = data_atual
dados_excel['Unidade'] = "MVME"

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_contratos = Aba_metricas['A3'].value
Valor_area_responsavel_contratos = Aba_metricas['B3'].value
Valor_setor_responsavel_contratos = Aba_metricas['C3'].value
Valor_frente2_contratos = Aba_metricas['D3'].value
Valor_indicador_contratos = Aba_metricas['E3'].value
Valor_meta_contratos = Aba_metricas['F3'].value
Valor_ruim_contratos = Aba_metricas['G3'].value
Valor_regular_contratos = Aba_metricas['H3'].value
Valor_otimo_contratos = Aba_metricas['I3'].value
Valor_peso_contratos = Aba_metricas['J3'].value

# Criar os dados para a tabela
dados_contrato = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_contratos,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_contratos,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_contratos,
    'FRENTE 2': Valor_frente2_contratos,
    'INDICADOR': Valor_indicador_contratos,
    'META': Valor_meta_contratos,
    'RUIM': Valor_ruim_contratos,
    'REGULAR': Valor_regular_contratos,
    'ÓTIMO': Valor_otimo_contratos,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_contratos,
}


# Criar o DataFrame
df = pd.DataFrame(dados_contrato)

# Inicializar as listas
Contratos = []

for mes in meses:
    if dados_excel['Mês'].isna().all():
        Contratos = None
    else:
        qtde_Contrato_Integral_total = dados_excel.loc[
            (dados_excel['Tipo (Integral, regular, futsal)'] == 'Integral') & (dados_excel['Mês'] == mes),
            'Total de contratos'].sum()
        qtde_Contrato_Regular_total = dados_excel.loc[
            (dados_excel['Tipo (Integral, regular, futsal)'] == 'Regular') & (dados_excel['Mês'] == mes),
            'Total de contratos'].sum()
        qtde_Contrato_Assinados_Integral_total = dados_excel.loc[
            (dados_excel['Tipo (Integral, regular, futsal)'] == 'Integral') & (dados_excel['Mês'] == mes),
            'Contratos assinados'].sum()
        qtde_Contrato_Assinados_Regular_total = dados_excel.loc[
            (dados_excel['Tipo (Integral, regular, futsal)'] == 'Regular') & (dados_excel['Mês'] == mes),
            'Contratos assinados'].sum()
        qtde_Contratos_Assinados_Total = qtde_Contrato_Assinados_Regular_total + qtde_Contrato_Assinados_Integral_total
        qtde_Contratos_Total = qtde_Contrato_Regular_total + qtde_Contrato_Integral_total

    if qtde_Contratos_Total>0:
        Contratos.append(qtde_Contratos_Assinados_Total / qtde_Contratos_Total)
    else:
        Contratos.append(None)

Contratos_Serie = pd.Series(Contratos)

Contratos_formatado = Contratos_Serie.map('{:.2%}'.format)

# Atribuir as listas ao DataFrame
df['Valor'] = Contratos_formatado

Tratamento_Contrato = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Contrato.xlsx'
df.to_excel(Tratamento_Contrato,index=False)

############################################# 7 - Colaboradores ##########################################################
############################################# Turnover
############################# INDICADOR : ((Pessoas contratadas + Pessoas Demitidas)/2)/Total de Colaboradores Ativos do período

caminho_arquivo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\7 - Colaboradores.xlsx'
nome_aba = '7 - Colaboradores'
dados_excel = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

# Obter a data atual
data_atual = datetime.today()

# Criar uma nova coluna com a data atual
dados_excel['Date_insert'] = data_atual
dados_excel['Unidade'] = "MVME"
dados_excel['Data de admissão'] = pd.to_datetime(dados_excel['Data de admissão'])
dados_excel['Mês_Admissão'] = dados_excel['Data de admissão'].dt.month
dados_excel['Ano_Admissão'] = dados_excel['Data de admissão'].dt.year
dados_excel['Data de demissão'] = pd.to_datetime(dados_excel['Data de demissão'])
dados_excel['Mês_demissão'] = dados_excel['Data de demissão'].dt.month
dados_excel['Ano_demissão'] = dados_excel['Data de demissão'].dt.year

# Obter o ano e mês atual
ano_atual = datetime.now().year
mes_atual = datetime.now().month

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))

# Inicializar as listas
qtde_total_colaboradores = []
qtde_admitidos_mes = []
qtde_demitidos_mes = []
qtde_demitidos_pedido = []

for mes in meses:
    # Contar a quantidade total de colaboradores
    if mes <= mes_atual:
        qtde_total_colaboradores.append(len(dados_excel[(dados_excel['Ano_Admissão'] < ano_atual) | ((dados_excel['Ano_Admissão'] == ano_atual) & (dados_excel['Mês_Admissão'] <= mes))]))
    else:
        qtde_total_colaboradores.append(0)

    # Contar a quantidade total de colaboradores admitidos no mês
    qtde_admitidos_mes.append(len(dados_excel[(dados_excel['Mês_Admissão'] == mes) & (dados_excel['Ano_Admissão'] == ano_atual)]))

    # Contar a quantidade total de colaboradores demitidos no mês
    qtde_demitidos_mes.append(len(dados_excel[(dados_excel['Mês_demissão'] == mes) & (dados_excel['Ano_demissão'] == ano_atual)]))

    # Contar a quantidade de colaboradores demitidos por pedido
    qtde_demitidos_pedido.append(len(dados_excel[(dados_excel['Mês_demissão'] == mes) & (dados_excel['Ano_demissão'] == ano_atual) & (dados_excel['Tipo de desligamento'] == 'Pedido de Demissão')]))


# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_turnover = Aba_metricas['A9'].value
Valor_area_responsavel_turnover = Aba_metricas['B9'].value
Valor_setor_responsavel_turnover = Aba_metricas['C9'].value
Valor_frente2_turnover = Aba_metricas['D9'].value
Valor_indicador_turnover = Aba_metricas['E9'].value
Valor_meta_turnover = Aba_metricas['F9'].value
Valor_ruim_turnover = Aba_metricas['G9'].value
Valor_regular_turnover = Aba_metricas['H9'].value
Valor_otimo_turnover = Aba_metricas['I9'].value
Valor_peso_turnover = Aba_metricas['J9'].value

# Criar o DataFrame de colaboradores
df_colaboradores = pd.DataFrame({
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_turnover,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_turnover,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_turnover,
    'FRENTE 2': Valor_frente2_turnover,
    'INDICADOR': Valor_indicador_turnover,
    'META': Valor_meta_turnover,
    'RUIM': Valor_ruim_turnover,
    'REGULAR': Valor_regular_turnover,
    'ÓTIMO': Valor_otimo_turnover,
    'Valor': [(qtde_admitidos_mes[i] + qtde_demitidos_mes[i]) / (2 * qtde_total_colaboradores[i]) if qtde_total_colaboradores[i] is not None and qtde_total_colaboradores[i] != 0 else None for i in range(len(meses))],
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_turnover,
})

df_colaboradores['Valor'] = df_colaboradores['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

# Salvar o DataFrame em um novo arquivo Excel
Tratamento_Turnover = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Turnover.xlsx'
df_colaboradores.to_excel(Tratamento_Turnover, index=False)

############################################# Turnover
############################# INDICADOR : Pessoas que pediram desligamento/Total de Colaboradores Ativos do período


# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_pedido_desligamento = Aba_metricas['A10'].value
Valor_area_responsavel_pedido_desligamento = Aba_metricas['B10'].value
Valor_setor_responsavel_pedido_desligamento = Aba_metricas['C10'].value
Valor_frente2_pedido_desligamento = Aba_metricas['D10'].value
Valor_indicador_pedido_desligamento = Aba_metricas['E10'].value
Valor_meta_pedido_desligamento = Aba_metricas['F10'].value
Valor_ruim_pedido_desligamento = Aba_metricas['G10'].value
Valor_regular_pedido_desligamento = Aba_metricas['H10'].value
Valor_otimo_pedido_desligamento = Aba_metricas['I10'].value
Valor_peso_pedido_desligamento = Aba_metricas['J10'].value


# Criar o DataFrame de colaboradores
df_colaboradores = pd.DataFrame({
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_pedido_desligamento,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_pedido_desligamento,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_pedido_desligamento,
    'FRENTE 2': Valor_frente2_pedido_desligamento,
    'INDICADOR': Valor_indicador_pedido_desligamento,
    'META': Valor_meta_pedido_desligamento,
    'RUIM': Valor_ruim_pedido_desligamento,
    'REGULAR': Valor_regular_pedido_desligamento,
    'ÓTIMO': Valor_otimo_pedido_desligamento,
    'Valor': [(qtde_admitidos_mes[i] + qtde_demitidos_mes[i]) / (2 * qtde_total_colaboradores[i]) if qtde_total_colaboradores[i] is not None and qtde_total_colaboradores[i] != 0 else 0 for i in range(len(meses))],
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_pedido_desligamento,
})

df_colaboradores['Valor'] = df_colaboradores['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

# Salvar o DataFrame em um novo arquivo Excel
Tratamento_Pedidos_desligamento = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_Pedidos_desligamento.xlsx'
df_colaboradores.to_excel(Tratamento_Pedidos_desligamento, index=False)


############################################# 14 - Notas_PH ##########################################################
############################################# Notas PH
############################# INDICADOR : Nº de alunos com média =>7 / Nº de alunos participantes

caminho_arquivo_ph = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\15 - Notas_PH.xlsx'
nome_aba_ph = 'PH'
dados_excel_ph = pd.read_excel(caminho_arquivo_ph, sheet_name=nome_aba_ph)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

Qtde_total_alunos_nota7_ph = dados_excel_ph.groupby('MÊS')['% RENDIMENTO DA TURMA'].mean()

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_notas_ph = Aba_metricas['A11'].value
Valor_area_responsavel_notas_ph = Aba_metricas['B11'].value
Valor_setor_responsavel_notas_ph = Aba_metricas['C11'].value
Valor_frente2_notas_ph = Aba_metricas['D11'].value
Valor_indicador_notas_ph = Aba_metricas['E11'].value
Valor_meta_notas_ph = Aba_metricas['F11'].value
Valor_ruim_notas_ph = Aba_metricas['G11'].value
Valor_regular_notas_ph = Aba_metricas['H11'].value
Valor_otimo_notas_ph = Aba_metricas['I11'].value
Valor_peso_notas_ph = Aba_metricas['J11'].value


# Criar o DataFrame de colaboradores
df_ph = pd.DataFrame({
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_notas_ph,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_notas_ph,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_notas_ph,
    'FRENTE 2': Valor_frente2_notas_ph,
    'INDICADOR': Valor_indicador_notas_ph,
    'META': Valor_meta_notas_ph,
    'RUIM': Valor_ruim_notas_ph,
    'REGULAR': Valor_ruim_notas_ph,
    'ÓTIMO': Valor_otimo_notas_ph,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_notas_ph,
})

# Inicializar as listas
qtde_alunos_ph = []


for mes in meses:
    if dados_excel_ph['MÊS'].isna().all():
        qtde_alunos_ph = None
    else:
        # Filtrar os dados para o mês atual
        dados_mes_ph = dados_excel_ph[dados_excel_ph['MÊS'] == mes]

        # Calcular a quantidade de alunos com nota maior que 7
        Qtde_total_alunos_nota7_ph = dados_mes_ph['% RENDIMENTO DA TURMA'].mean()

        # Adicionar à lista de quantidade de alunos
        qtde_alunos_ph.append(Qtde_total_alunos_nota7_ph)

df_ph['Valor'] = qtde_alunos_ph
df_ph['Valor'] = df_ph['Valor'].map('{:.2%}'.format)

Tratamento_ph = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_ph.xlsx'
df_ph.to_excel(Tratamento_ph, index=False)

############################################# Engajamento
############################# INDICADOR : Nº de alunos participantes da prova/ Nº alunos elegíveis

caminho_arquivo_engajamento_ph = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\14 - Engajamento_PH.xlsx'
nome_aba_engajamento_ph = 'PH'
dados_excel_ph = pd.read_excel(caminho_arquivo_engajamento_ph, sheet_name=nome_aba_engajamento_ph)


# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_engajamento_ph = Aba_metricas['A12'].value
Valor_area_responsavel_engajamento_ph = Aba_metricas['B12'].value
Valor_setor_responsavel_engajamento_ph = Aba_metricas['C12'].value
Valor_frente2_engajamento_ph = Aba_metricas['D12'].value
Valor_indicador_engajamento_ph = Aba_metricas['E12'].value
Valor_meta_engajamento_ph = Aba_metricas['F12'].value
Valor_ruim_engajamento_ph = Aba_metricas['G12'].value
Valor_regular_engajamento_ph = Aba_metricas['H12'].value
Valor_otimo_engajamento_ph = Aba_metricas['I12'].value
Valor_peso_engajamento_ph = Aba_metricas['J12'].value


# Criar o DataFrame de colaboradores
df_avaliacoes = pd.DataFrame({
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_engajamento_ph,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_engajamento_ph,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_engajamento_ph,
    'FRENTE 2': Valor_frente2_engajamento_ph,
    'INDICADOR': Valor_indicador_engajamento_ph,
    'META': Valor_meta_engajamento_ph,
    'RUIM': Valor_ruim_engajamento_ph,
    'REGULAR': Valor_regular_engajamento_ph,
    'ÓTIMO': Valor_otimo_engajamento_ph,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_engajamento_ph,
})

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Inicializar as listas
qtde_engajamento = []

for mes in meses:
    if dados_excel_ph['MÊS'].isna().all():
        qtde_engajamento = None
    else:
        # Filtrar os dados para o mês atual
        dados_mes_ph = dados_excel_ph[dados_excel_ph['MÊS'] == mes]

        # Calcular a quantidade total de alunos participantes da prova
        qtde_alunos_participantes = dados_mes_ph['% ENGAJAMENTO'].mean()

        engajamento = qtde_alunos_participantes
        # Adicionar o valor de engajamento à lista
        qtde_engajamento.append(engajamento)

# Atualizar o valor no DataFrame de avaliações
df_avaliacoes['Valor'] = qtde_engajamento
df_avaliacoes['Valor'] = df_avaliacoes['Valor'].map('{:.2%}'.format)

# Salvar o DataFrame em um novo arquivo Excel
Tratamento_engajamento = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_engajamento.xlsx'
df_avaliacoes.to_excel(Tratamento_engajamento, index=False)

############################################# 13 - Atrasos_Faltas_alunos ##########################################################
############################################# Tabela Falta aluno
############################# INDICADOR : qtd de alunos que tiveram pelo menos 1 falta no período/ qtd de alunos matriculados

caminho_arquivo_falta_aluno = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13 - Faltas_alunos.xlsx'
nome_aba_falta_aluno = 'ABSENTEISMO_PONTUALIDADE'
dados_excel_falta_aluno = pd.read_excel(caminho_arquivo_falta_aluno, sheet_name=nome_aba_falta_aluno)

########################################### ALUNOS TOTAL ################################

caminho_arquivo_alunos_total = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\11 - Listagem_dos_alunos.xlsx'
nome_aba_alunos_total = '15 - Listagem dos alunos - atua'
dados_excel_alunos_total = pd.read_excel(caminho_arquivo_alunos_total, sheet_name=nome_aba_alunos_total)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_falta_aluno = Aba_metricas['A13'].value
Valor_area_responsavel_falta_aluno = Aba_metricas['B13'].value
Valor_setor_responsavel_falta_aluno = Aba_metricas['C13'].value
Valor_frente2_falta_aluno = Aba_metricas['D13'].value
Valor_indicador_falta_aluno = Aba_metricas['E13'].value
Valor_meta_falta_aluno = Aba_metricas['F13'].value
Valor_ruim_falta_aluno = Aba_metricas['G13'].value
Valor_regular_falta_aluno = Aba_metricas['H13'].value
Valor_otimo_falta_aluno = Aba_metricas['I13'].value
Valor_peso_falta_aluno = Aba_metricas['J13'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_falta_aluno,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_falta_aluno,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_falta_aluno,
    'FRENTE 2': Valor_frente2_falta_aluno,
    'INDICADOR': Valor_indicador_falta_aluno,
    'META': Valor_meta_falta_aluno,
    'RUIM': Valor_ruim_falta_aluno,
    'REGULAR': Valor_regular_falta_aluno,
    'ÓTIMO': Valor_otimo_falta_aluno,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_falta_aluno,
}

# Criar o DataFrame
df_falta_aluno = pd.DataFrame(dados)

# Inicializar as listas
Qtde_Valor_falta_aluno = []
# Calcular o total de alunos na disciplina
Qtde_total_alunos_disciplina = dados_excel_alunos_total['Numero de Matrícula'].count()

for mes in meses:
    dados_mes = dados_excel_falta_aluno.loc[dados_excel_falta_aluno['MÊS'] == mes]

    if dados_excel_falta_aluno['QTDE DIAS DE FALTA'].isna().all():
        taxa_falta_aluno = None
    else:
        qtde_falta_aluno_total = dados_excel_falta_aluno.loc[dados_excel_falta_aluno['MÊS'] == mes, 'QTDE DIAS DE FALTA'].sum()
        if qtde_falta_aluno_total == 0:
            taxa_falta_aluno = None
        else:
            if Qtde_total_alunos_disciplina != 0:
                taxa_falta_aluno = qtde_falta_aluno_total / Qtde_total_alunos_disciplina
            else:
                taxa_falta_aluno = None

    Qtde_Valor_falta_aluno.append(taxa_falta_aluno)

# Atribuir as listas ao DataFrame
df_falta_aluno['Valor'] = Qtde_Valor_falta_aluno

Tratamento_falta_aluno = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_falta_aluno.xlsx'
df_falta_aluno.to_excel(Tratamento_falta_aluno,index=False)

############################################# 10 - Ensalamento ##########################################################
############################################# Tabela Ensalamento
############################# INDICADOR : alunos na turma/x (x =capacidade limitante da turma (capacidade de infraestrutura ou capacidade pedagógica)

caminho_arquivo_ensalamento = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\10 - Ensalamento.xlsx'
nome_aba_ensalamento = '10 - Ensalamento'
dados_excel_ensalamento = pd.read_excel(caminho_arquivo_ensalamento, sheet_name=nome_aba_ensalamento)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year
mes_atual = datetime.now().month

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_ensalamento = Aba_metricas['A14'].value
Valor_area_responsavel_ensalamento = Aba_metricas['B14'].value
Valor_setor_responsavel_ensalamento = Aba_metricas['C14'].value
Valor_frente2_ensalamento = Aba_metricas['D14'].value
Valor_indicador_ensalamento = Aba_metricas['E14'].value
Valor_meta_ensalamento = Aba_metricas['F14'].value
Valor_ruim_ensalamento = Aba_metricas['G14'].value
Valor_regular_ensalamento = Aba_metricas['H14'].value
Valor_otimo_ensalamento = Aba_metricas['I14'].value
Valor_peso_ensalamento = Aba_metricas['J14'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_ensalamento,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_ensalamento,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_ensalamento,
    'FRENTE 2': Valor_frente2_ensalamento,
    'INDICADOR': Valor_indicador_ensalamento,
    'META': Valor_meta_ensalamento,
    'RUIM': Valor_ruim_ensalamento,
    'REGULAR': Valor_regular_ensalamento,
    'ÓTIMO': Valor_otimo_ensalamento,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_ensalamento,
}

# Criar o DataFrame
df_ensalamento = pd.DataFrame(dados)

# Inicializar as listas
Qtde_ensalamento = []

# Inicializar as listas
media_ensalamento_por_mes = []

for mes in meses:
    ensalamentos_mes = []
    if dados_excel_ensalamento['Capacidade pedagógica'].isna().all():
        media_ensalamento_mes = None
    else:
        for _, row in dados_excel_ensalamento[dados_excel_ensalamento['MÊS'] == mes].iterrows():
            if row['Capacidade Alunos (Total)'] > 0:  # Verificar se a quantidade de alunos é maior que zero
                if row['CONSIDERAR'] == 'PED':
                    ensalamento = row['Capacidade Alunos (Total)'] / row['Capacidade pedagógica']
                elif row['CONSIDERAR'] == '':
                    ensalamento = 0  # Não considerar na média
                else:
                    ensalamento = row['Capacidade Alunos (Total)'] / row['Capacidade Infraestrutura']

                ensalamentos_mes.append(ensalamento)

    if ensalamentos_mes:
        media_ensalamento_mes = sum(ensalamentos_mes) / len(ensalamentos_mes)
    else:
        media_ensalamento_mes = None

    media_ensalamento_por_mes.append(media_ensalamento_mes)

# Preencher a coluna 'Valor' com as médias de ensalamento por mês
df_ensalamento['Valor'] = media_ensalamento_por_mes

df_ensalamento['Valor'] = df_ensalamento['Valor'].map('{:.2%}'.format)

Tratamento_ensalamento = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_ensalamento.xlsx'
df_ensalamento.to_excel(Tratamento_ensalamento,index=False)

############################################# 22 - Governanca ##########################################################
############################################# Tabela Governança
############################# INDICADOR : nº de reuniões do mês realizadas pelas governança escolar/ nº reuniões de governança escolar exigidas

caminho_arquivo_governanca = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\22 - Governanca.xlsx'
nome_aba_governanca = 'CM'
dados_excel_governanca = pd.read_excel(caminho_arquivo_governanca, sheet_name=nome_aba_governanca)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_governanca = Aba_metricas['A15'].value
Valor_area_responsavel_governanca = Aba_metricas['B15'].value
Valor_setor_responsavel_governanca = Aba_metricas['C15'].value
Valor_frente2_governanca = Aba_metricas['D15'].value
Valor_indicador_governanca = Aba_metricas['E15'].value
Valor_meta_governanca = Aba_metricas['F15'].value
Valor_ruim_governanca = Aba_metricas['G15'].value
Valor_regular_governanca = Aba_metricas['H15'].value
Valor_otimo_governanca = Aba_metricas['I15'].value
Valor_governanca = Aba_metricas['J15'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_governanca,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_governanca,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_governanca,
    'FRENTE 2': Valor_frente2_governanca,
    'INDICADOR': Valor_indicador_governanca,
    'META': Valor_meta_governanca,
    'RUIM': Valor_ruim_governanca,
    'REGULAR': Valor_regular_governanca,
    'ÓTIMO': Valor_otimo_governanca,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_governanca,
}

# Criar o DataFrame
df_governanca = pd.DataFrame(dados)

Qtde_reunioes = []

for mes in meses:
    if dados_excel_governanca['Qtde realizada'].isna().all():
        Qtde_reunioes.append(None)
    else:
        qtde_reunioes_exigidas = dados_excel_governanca.loc[dados_excel_governanca['MÊS'] == mes, 'Qtde de reuniões programadas'].sum()
        qtde_reunioes_realizadas = dados_excel_governanca.loc[dados_excel_governanca['MÊS'] == mes, 'Qtde realizada'].sum()

        if qtde_reunioes_realizadas == 0:
            Qtde_reunioes.append(None)
        else:
            Qtde_reunioes.append((qtde_reunioes_realizadas/qtde_reunioes_exigidas))

df_governanca['Valor'] = Qtde_reunioes

df_governanca['Valor'] = df_governanca['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_governanca = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_governanca.xlsx'
df_governanca.to_excel(Tratamento_governanca,index=False)

############################################# 12 - Aderencia_Totvs ##########################################################
############################################# Tabela Aderencia Totvs
############################# INDICADOR : Aderência do Totvs

caminho_arquivo_totvs = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\12 - Aderencia_Totvs.xlsx'
nome_aba_totvs = '16 - Aderência Totvs'
dados_excel_totvs = pd.read_excel(caminho_arquivo_totvs, sheet_name=nome_aba_totvs)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_aderencia_totvs = Aba_metricas['A16'].value
Valor_area_responsavel_aderencia_totvs = Aba_metricas['B16'].value
Valor_setor_responsavel_aderencia_totvs = Aba_metricas['C16'].value
Valor_frente2_aderencia_totvs = Aba_metricas['D16'].value
Valor_indicador_aderencia_totvs = Aba_metricas['E16'].value
Valor_meta_aderencia_totvs = Aba_metricas['F16'].value
Valor_ruim_aderencia_totvs = Aba_metricas['G16'].value
Valor_regular_aderencia_totvs = Aba_metricas['H16'].value
Valor_otimo_aderencia_totvs = Aba_metricas['I16'].value
Valor_peso_aderencia_totvs = Aba_metricas['J16'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_aderencia_totvs,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_aderencia_totvs,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_aderencia_totvs,
    'FRENTE 2': Valor_frente2_aderencia_totvs,
    'INDICADOR': Valor_indicador_aderencia_totvs,
    'META': Valor_meta_aderencia_totvs,
    'RUIM': Valor_ruim_aderencia_totvs,
    'REGULAR': Valor_regular_aderencia_totvs,
    'ÓTIMO': Valor_otimo_aderencia_totvs,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_aderencia_totvs,
}

# Criar o DataFrame
df_totvs = pd.DataFrame(dados)
qtde_implantada = 0

# Inicializar as contagens
qtde_implantada = 0
qtde_total_etapas = 0

# Iterar sobre as linhas do DataFrame 'dados_excel_totvs'
for index, row in dados_excel_totvs.iterrows():
    if row['STATUS'] == 'Implantado':
        qtde_implantada += 1
        qtde_total_etapas += 1
    else:
        qtde_total_etapas += 1

# Calcular a proporção de etapas concluídas
proporcao_concluida = qtde_implantada / qtde_total_etapas

df_totvs['Valor'] = proporcao_concluida

df_totvs['Valor'] = df_totvs['Valor'].map('{:.2%}'.format)

Tratamento_totvs = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_totvs.xlsx'
df_totvs.to_excel(Tratamento_totvs,index=False)

############################################# 20 - Financeiro_Receita ##########################################################
############################################# Receita

caminho_arquivo_receita = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.3. Financeiro - MVM\20 - Financeiro_Receita.xlsx'
nome_aba_receita = 'Receita'
dados_excel_receita = pd.read_excel(caminho_arquivo_receita, sheet_name=nome_aba_receita)



############################################# 18 - Financeiro_Inadimplencia ##########################################################
############################################# Tabela de Alunos inadimplentes
############################# INDICADOR : alunos inadimplentes/alunos total

caminho_arquivo_inadimplencia = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.3. Financeiro - MVM\18 - Financeiro_Inadimplencia.xlsx'
nome_aba_inadimplencia = 'INADIMPLENCIA'
dados_excel_inadimplencia = pd.read_excel(caminho_arquivo_inadimplencia, sheet_name=nome_aba_inadimplencia)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_inadimplencia = Aba_metricas['A8'].value
Valor_area_responsavel_financeiro_inadimplencia = Aba_metricas['B8'].value
Valor_setor_responsavel_financeiro_inadimplencia = Aba_metricas['C8'].value
Valor_frente2_financeiro_inadimplencia = Aba_metricas['D8'].value
Valor_indicador_financeiro_inadimplencia = Aba_metricas['E8'].value
Valor_meta_financeiro_inadimplencia = Aba_metricas['F8'].value
Valor_ruim_financeiro_inadimplencia = Aba_metricas['G8'].value
Valor_regular_financeiro_inadimplencia = Aba_metricas['H8'].value
Valor_otimo_financeiro_inadimplencia = Aba_metricas['I8'].value
Valor_peso_inadimplencia = Aba_metricas['J8'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_inadimplencia,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_financeiro_inadimplencia,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_financeiro_inadimplencia,
    'FRENTE 2': Valor_frente2_financeiro_inadimplencia,
    'INDICADOR': Valor_indicador_financeiro_inadimplencia,
    'META': Valor_meta_financeiro_inadimplencia,
    'RUIM': Valor_ruim_financeiro_inadimplencia,
    'REGULAR': Valor_regular_financeiro_inadimplencia,
    'ÓTIMO': Valor_otimo_financeiro_inadimplencia,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_inadimplencia,
}

# Criar o DataFrame
df_financeiro_inadimplencia = pd.DataFrame(dados)
prop_alunos_inadimplentes = []

for mes in meses:
    if dados_excel_inadimplencia['Mês'].isna().all():
        qtde_valor_inadimplentes = None
        qtde_valor_receita_inadimplentes = None
    else:
        # Filtrar os dados para o mês atual
        dados_mes_inadimplencia = dados_excel_inadimplencia[dados_excel_inadimplencia['Mês'] == mes]
        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_inadimplentes = dados_mes_inadimplencia['Valor Inadimplência'].sum()

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_inadimplentes = dados_mes_receita['Valor'].sum()

    try:
        if qtde_valor_inadimplentes is None or qtde_valor_receita_inadimplentes is None:
            prop_alunos_inadimplentes.append(None)
        else:
            prop_alunos_inadimplentes.append(qtde_valor_inadimplentes / qtde_valor_receita_inadimplentes)
    except ZeroDivisionError:
        prop_alunos_inadimplentes.append(None)

df_financeiro_inadimplencia['Valor'] = prop_alunos_inadimplentes

df_financeiro_inadimplencia['Valor'] = df_financeiro_inadimplencia['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_alunos_inadimplentes = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_alunos_inadimplentes.xlsx'
df_financeiro_inadimplencia.to_excel(Tratamento_alunos_inadimplentes,index=False)

############################################# 17 - Financeiro_Custos ##########################################################
############################################# Tabela de Custo por aluno
############################# INDICADOR : Outros Custos/receita

caminho_arquivo_custos = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.3. Financeiro - MVM\17 - Financeiro_Custos.xlsx'
nome_aba_custos = 'CUSTOS'
dados_excel_custos = pd.read_excel(caminho_arquivo_custos, sheet_name=nome_aba_custos)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_custo_aluno = Aba_metricas['A6'].value
Valor_area_responsavel_financeiro_custo_aluno = Aba_metricas['B6'].value
Valor_setor_responsavel_financeiro_custo_aluno = Aba_metricas['C6'].value
Valor_frente2_financeiro_custo_aluno = Aba_metricas['D6'].value
Valor_indicador_financeiro_custo_aluno = Aba_metricas['E6'].value
Valor_meta_financeiro_custo_aluno = Aba_metricas['F6'].value
Valor_ruim_financeiro_custo_aluno = Aba_metricas['G6'].value
Valor_regular_financeiro_custo_aluno = Aba_metricas['H6'].value
Valor_otimo_financeiro_custo_aluno = Aba_metricas['I6'].value
Valor_peso_custo_aluno = Aba_metricas['J6'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_custo_aluno,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_financeiro_custo_aluno,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_financeiro_custo_aluno,
    'FRENTE 2': Valor_frente2_financeiro_custo_aluno,
    'INDICADOR': Valor_indicador_financeiro_custo_aluno,
    'META': Valor_meta_financeiro_custo_aluno,
    'RUIM': Valor_ruim_financeiro_custo_aluno,
    'REGULAR': Valor_regular_financeiro_custo_aluno,
    'ÓTIMO': Valor_otimo_financeiro_custo_aluno,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_custo_aluno,
}

# Criar o DataFrame
df_financeiro_custos = pd.DataFrame(dados)
custos_alunos_total = []

for mes in meses:
    if dados_excel_custos['Mês'].isna().all():
        qtde_outros_custos = None
        qtde_valor_receita_outros_custos = None
    else:
        # Filtrar os dados para o mês atual
        dados_custos = dados_excel_custos[
            (dados_excel_custos['Mês'] == mes) & (dados_excel_custos['Custo'] == 'Outros Custos')]

        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_outros_custos = dados_mes_receita['Valor'].sum()

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_outros_custos = dados_custos['Valor'].sum()

    try:
        if qtde_outros_custos is None or qtde_valor_receita_outros_custos is None:
            custos_alunos_total.append (None)
        else:
            custos_alunos_total.append(qtde_outros_custos / qtde_valor_receita_outros_custos)
    except ZeroDivisionError:
        custos_alunos_total.append(None)

df_financeiro_custos['Valor'] = custos_alunos_total

df_financeiro_custos['Valor'] = df_financeiro_custos['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_custos = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_custos.xlsx'
df_financeiro_custos.to_excel(Tratamento_custos,index=False)

############################################# 21 - Financeiro_Orcamento ##########################################################
############################################# Tabela de Orçado/Realizado
############################# INDICADOR : Valor orçado/ Valor realizado

caminho_arquivo_orcado = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.3. Financeiro - MVM\21 - Financeiro_Orcamento.xlsx'
nome_aba_orcado = 'ORÇAMENTO'
dados_excel_orcado = pd.read_excel(caminho_arquivo_orcado, sheet_name=nome_aba_orcado)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_orcamento = Aba_metricas['A9'].value
Valor_area_responsavel_financeiro_orcamento = Aba_metricas['B9'].value
Valor_setor_responsavel_financeiro_orcamento = Aba_metricas['C9'].value
Valor_frente2_financeiro_orcamento = Aba_metricas['D9'].value
Valor_indicador_financeiro_orcamento = Aba_metricas['E9'].value
Valor_meta_financeiro_orcamento = Aba_metricas['F9'].value
Valor_ruim_financeiro_orcamento = Aba_metricas['G9'].value
Valor_regular_financeiro_orcamento = Aba_metricas['H9'].value
Valor_otimo_financeiro_orcamento = Aba_metricas['I9'].value
Valor_peso_orcamento = Aba_metricas['J9'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_orcamento,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_financeiro_orcamento,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_financeiro_orcamento,
    'FRENTE 2': Valor_frente2_financeiro_orcamento,
    'INDICADOR': Valor_indicador_financeiro_orcamento,
    'META': Valor_meta_financeiro_orcamento,
    'RUIM': Valor_ruim_financeiro_orcamento,
    'REGULAR': Valor_regular_financeiro_orcamento,
    'ÓTIMO': Valor_otimo_financeiro_orcamento,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_orcamento,
}

# Criar o DataFrame
df_financeiro_orcado_realizado = pd.DataFrame(dados)
custos_orcado = []

for mes in meses:
    if dados_excel_orcado['MÊS'].isna().all():
        custos_orcado = None
    else:
        # Filtrar os dados para o mês atual
        dados_orcado = dados_excel_orcado[dados_excel_orcado['MÊS'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        valor_orcado = dados_orcado['ORÇADO'].sum()
        valor_realizado = dados_orcado['REALIZADO'].sum()

    try:
        custos_orcado.append(valor_realizado / valor_orcado)
    except (ZeroDivisionError, RuntimeWarning):
        custos_orcado.append(None)

df_financeiro_orcado_realizado['Valor'] = custos_orcado

df_financeiro_orcado_realizado['Valor'] = df_financeiro_orcado_realizado['Valor'].map('{:.2%}'.format)

Tratamento_orcado_realizado = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_orcado_realizado.xlsx'
df_financeiro_orcado_realizado.to_excel(Tratamento_orcado_realizado,index=False)

############################################# 17 - Financeiro_Custos ##########################################################
############################################# Tabela Custos pedagógicos por aluno
############################# INDICADOR : gasto folha pedagógica/ nº alunos

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_folha_pedagogico = Aba_metricas['A3'].value
Valor_area_responsavel_folha_pedagogico = Aba_metricas['B3'].value
Valor_setor_responsavel_folha_pedagogico = Aba_metricas['C3'].value
Valor_frente2_folha_pedagogico = Aba_metricas['D3'].value
Valor_indicador_folha_pedagogico = Aba_metricas['E3'].value
Valor_meta_folha_pedagogico = Aba_metricas['F3'].value
Valor_ruim_folha_pedagogico = Aba_metricas['G3'].value
Valor_regular_folha_pedagogico = Aba_metricas['H3'].value
Valor_otimo_folha_pedagogico = Aba_metricas['I3'].value
Valor_peso_folha_pedagogico = Aba_metricas['J3'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_folha_pedagogico,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_folha_pedagogico,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_folha_pedagogico,
    'FRENTE 2': Valor_frente2_folha_pedagogico,
    'INDICADOR': Valor_indicador_folha_pedagogico,
    'META': Valor_meta_folha_pedagogico,
    'RUIM': Valor_ruim_folha_pedagogico,
    'REGULAR': Valor_regular_folha_pedagogico,
    'ÓTIMO': Valor_otimo_folha_pedagogico,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_folha_pedagogico
}

# Criar o DataFrame
df_financeiro_custos_pedagogico = pd.DataFrame(dados)
custos_pedagogico = []

for mes in meses:
    if dados_excel_custos['Mês'].isna().all():
        custo_pedagogico = None
        qtde_valor_receita_folha_pedagogica = None
    else:
        # Filtrar os dados para o mês atual e para o custo "Pessoal Pedagógico"
        dados_custos_pedagogico = dados_excel_custos[(dados_excel_custos['Mês'] == mes) & (dados_excel_custos['Custo'] == 'Pessoal Pedagógico')]

        # Obter o valor do custo "Pessoal Pedagógico" para o mês atual
        custo_pedagogico = dados_custos_pedagogico['Valor'].sum()

        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_folha_pedagogica = dados_mes_receita['Valor'].sum()

    try:
        if custo_pedagogico is None or qtde_valor_receita_folha_pedagogica is None:
            custos_pedagogico.append(None)
        else:
            custos_pedagogico.append(custo_pedagogico / qtde_valor_receita_folha_pedagogica)
    except ZeroDivisionError:
        custos_pedagogico.append(None)

df_financeiro_custos_pedagogico['Valor'] = custos_pedagogico

df_financeiro_custos_pedagogico['Valor'] = df_financeiro_custos_pedagogico['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_custos_pedagogico = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_custos_pedagogico.xlsx'
df_financeiro_custos_pedagogico.to_excel(Tratamento_custos_pedagogico,index=False)

############################################# 17 - Financeiro_Custos ##########################################################
############################################# Tabela Custos Despesa administrativo por aluno
############################# INDICADOR : gasto folha adm/ nº alunos

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_folha_administrativo = Aba_metricas['A4'].value
Valor_area_responsavel_folha_administrativo = Aba_metricas['B4'].value
Valor_setor_responsavel_folha_administrativo = Aba_metricas['C4'].value
Valor_frente2_folha_administrativo = Aba_metricas['D4'].value
Valor_indicador_folha_administrativo = Aba_metricas['E4'].value
Valor_meta_folha_administrativo = Aba_metricas['F4'].value
Valor_ruim_folha_administrativo = Aba_metricas['G4'].value
Valor_regular_folha_administrativo= Aba_metricas['H4'].value
Valor_otimo_folha_administrativo = Aba_metricas['I4'].value
Valor_peso_folha_administrativo = Aba_metricas['J4'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_folha_administrativo,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_folha_administrativo,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_folha_administrativo,
    'FRENTE 2': Valor_frente2_folha_administrativo,
    'INDICADOR': Valor_indicador_folha_administrativo,
    'META': Valor_meta_folha_administrativo,
    'RUIM': Valor_ruim_folha_administrativo,
    'REGULAR': Valor_regular_folha_administrativo,
    'ÓTIMO': Valor_otimo_folha_administrativo,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_folha_administrativo,
}

# Criar o DataFrame
df_financeiro_custos_administrativo = pd.DataFrame(dados)
custos_administrativo = []

for mes in meses:
    if dados_excel_custos['Mês'].isna().all():
        custo_administrativo = None
        qtde_valor_receita_folha_administrativa = None
    else:
        # Filtrar os dados para o mês atual e para o custo "Pessoal Pedagógico"
        dados_custos_administrativo = dados_excel_custos[(dados_excel_custos['Mês'] == mes) & (dados_excel_custos['Custo'] == 'Pessoal Administrativo')]

        # Obter o valor do custo "Pessoal Pedagógico" para o mês atual
        custo_administrativo = dados_custos_administrativo['Valor'].sum()

        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_folha_administrativa = dados_mes_receita['Valor'].sum()

    try:
        if custo_administrativo is None or qtde_valor_receita_folha_administrativa is None:
            custos_administrativo.append(None)
        else:
            custos_administrativo.append(custo_administrativo / qtde_valor_receita_folha_administrativa)
    except ZeroDivisionError:
        custos_administrativo.append(None)

df_financeiro_custos_administrativo['Valor'] = custos_administrativo

df_financeiro_custos_administrativo['Valor'] = df_financeiro_custos_administrativo['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_custos_administrativo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_custos_administrativo.xlsx'
df_financeiro_custos_administrativo.to_excel(Tratamento_custos_administrativo,index=False)

############################################# 17 - Financeiro_Custos ##########################################################
############################################# Tabela Custos não pedagógicos por aluno
############################# INDICADOR : despesas administrativa/ nº alunos

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_despesa_administrativo = Aba_metricas['A5'].value
Valor_area_responsavel_despesa_administrativo = Aba_metricas['B5'].value
Valor_setor_responsavel_despesa_administrativo = Aba_metricas['C5'].value
Valor_frente2_despesa_administrativo = Aba_metricas['D5'].value
Valor_indicador_despesa_administrativo = Aba_metricas['E5'].value
Valor_meta_despesa_administrativo = Aba_metricas['F5'].value
Valor_ruim_despesa_administrativo = Aba_metricas['G5'].value
Valor_regular_despesa_administrativo= Aba_metricas['H5'].value
Valor_otimo_despesa_administrativo = Aba_metricas['I5'].value
Valor_peso_despesa_administrativo = Aba_metricas['J5'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_despesa_administrativo,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_despesa_administrativo,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_despesa_administrativo,
    'FRENTE 2': Valor_frente2_despesa_administrativo,
    'INDICADOR': Valor_indicador_despesa_administrativo,
    'META': Valor_meta_despesa_administrativo,
    'RUIM': Valor_ruim_despesa_administrativo,
    'REGULAR': Valor_regular_despesa_administrativo,
    'ÓTIMO': Valor_otimo_despesa_administrativo,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_despesa_administrativo,
}

# Criar o DataFrame
df_financeiro_despesa_administrativo = pd.DataFrame(dados)
custos_despesa_administrativo = []

for mes in meses:
    if dados_excel_custos['Mês'].isna().all():
        despesa_administrativo = None
        qtde_valor_receita_despesa_administrativa = None
    else:
        # Filtrar os dados para o mês atual e para o custo "Pessoal Pedagógico"
        dados_despesa_administrativo = dados_excel_custos[(dados_excel_custos['Mês'] == mes) & (dados_excel_custos['Custo'] == 'Despesas administrativas')]

        # Obter o valor do custo "Pessoal Pedagógico" para o mês atual
        despesa_administrativo = dados_despesa_administrativo['Valor'].sum()

        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_despesa_administrativa = dados_mes_receita['Valor'].sum()

    try:
        if despesa_administrativo is None or qtde_valor_receita_despesa_administrativa is None:
            custos_despesa_administrativo.append(None)
        else:
            custos_despesa_administrativo.append(despesa_administrativo / qtde_valor_receita_despesa_administrativa)
    except ZeroDivisionError:
        custos_despesa_administrativo.append(None)


df_financeiro_despesa_administrativo['Valor'] = custos_despesa_administrativo

# Formatar a coluna "Valor" como moeda brasileira (R$)
df_financeiro_despesa_administrativo['Valor'] = df_financeiro_despesa_administrativo['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_despesa_administrativo = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_despesa_administrativo.xlsx'
df_financeiro_despesa_administrativo.to_excel(Tratamento_despesa_administrativo,index=False)

############################################# 17 - Financeiro_Custos ##########################################################
############################################# Tabela Custos não pedagógicos por aluno
############################# INDICADOR : outras despesas administrativa/ nº alunos

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_financeiro_despesa_outras = Aba_metricas['A7'].value
Valor_area_responsavel_despesa_outras = Aba_metricas['B7'].value
Valor_setor_responsavel_despesa_outras = Aba_metricas['C7'].value
Valor_frente2_despesa_outras = Aba_metricas['D7'].value
Valor_indicador_despesa_outras = Aba_metricas['E7'].value
Valor_meta_despesa_outras = Aba_metricas['F7'].value
Valor_ruim_despesa_outras = Aba_metricas['G7'].value
Valor_regular_despesa_outras = Aba_metricas['H7'].value
Valor_otimo_despesa_outras = Aba_metricas['I7'].value
Valor_peso_despesa_outras = Aba_metricas['J7'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_financeiro_despesa_outras,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_despesa_outras,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_despesa_outras,
    'FRENTE 2': Valor_frente2_despesa_outras,
    'INDICADOR': Valor_indicador_despesa_outras,
    'META': Valor_meta_despesa_outras,
    'RUIM': Valor_ruim_despesa_outras,
    'REGULAR': Valor_regular_despesa_outras,
    'ÓTIMO': Valor_otimo_despesa_outras,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_despesa_outras,
}

# Criar o DataFrame
df_financeiro_despesa_outras = pd.DataFrame(dados)
custos_despesa_outras = []

for mes in meses:
    if dados_excel_custos['Mês'].isna().all():
        despesa_outras = None
        qtde_valor_receita_despesa_outras = None
    else:
        # Filtrar os dados para o mês atual e para o custo "Pessoal Pedagógico"
        dados_despesa_outras = dados_excel_custos[(dados_excel_custos['Mês'] == mes) & (dados_excel_custos['Custo'] == 'Outras despesas')]

        # Obter o valor do custo "Pessoal Pedagógico" para o mês atual
        despesa_outras = dados_despesa_outras['Valor'].sum()

        # Filtrar os dados para o mês atual
        dados_mes_receita = dados_excel_receita[dados_excel_receita['Mês'] == mes]

        # Obter a quantidade de alunos inadimplentes para o mês atual
        qtde_valor_receita_despesa_outras = dados_mes_receita['Valor'].sum()
    try:
        if despesa_outras is None or qtde_valor_receita_despesa_outras is None:
            custos_despesa_outras.append(None)
        else:
            custos_despesa_outras.append(despesa_outras / qtde_valor_receita_despesa_outras)
    except ZeroDivisionError:
        custos_despesa_outras.append(None)

df_financeiro_despesa_outras['Valor'] = custos_despesa_outras

# Formatar a coluna "Valor" como moeda brasileira (R$)
df_financeiro_despesa_outras['Valor'] = df_financeiro_despesa_outras['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_despesa_outras = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_despesa_outras.xlsx'
df_financeiro_despesa_outras.to_excel(Tratamento_despesa_outras,index=False)


############################################# 9 - Infraestrutura ##########################################################
############################################# Não acompanhado

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_infraestrutura = Aba_metricas['A17'].value
Valor_area_responsavel_infraestrutura = Aba_metricas['B17'].value
Valor_setor_responsavel_infraestrutura = Aba_metricas['C17'].value
Valor_frente2_infraestrutura = Aba_metricas['D17'].value
Valor_indicador_infraestrutura = Aba_metricas['E17'].value
Valor_meta_infraestrutura = Aba_metricas['F17'].value
Valor_ruim_infraestrutura = Aba_metricas['G17'].value
Valor_regular_infraestrutura = Aba_metricas['H17'].value
Valor_otimo_infraestrutura = Aba_metricas['I17'].value
Valor_peso_infraestrutra = Aba_metricas['J17'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_infraestrutura,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_infraestrutura,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_infraestrutura,
    'FRENTE 2': Valor_frente2_infraestrutura,
    'INDICADOR': Valor_indicador_infraestrutura,
    'META': Valor_meta_infraestrutura,
    'RUIM': Valor_ruim_infraestrutura,
    'REGULAR': Valor_regular_infraestrutura,
    'ÓTIMO': Valor_otimo_infraestrutura,
    'Valor': [0] * len(meses),
    'Tipo':["a definir"] * len(meses),
    'Peso': Valor_peso_infraestrutra,
}

# Criar o DataFrame
df_infraestrutura = pd.DataFrame(dados)
infraestrutura = []

Tratamento_infraestrutura = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_infraestrutura.xlsx'
df_infraestrutura.to_excel(Tratamento_infraestrutura,index=False)

############################################# NPS COLABORADORES ##########################################################
############################# INDICADOR : Resultado da nota NPS Colaboradores aplicado nas escolas

caminho_arquivo_nps_colaboradores = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.4. NPS Colaboradores\24. NPS Colaboradores.xlsx'
nome_aba_nps_colaboradores = 'Planilha1'
dados_excel_nps_colaboradores = pd.read_excel(caminho_arquivo_nps_colaboradores, sheet_name=nome_aba_nps_colaboradores)


# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_nps_colaboradores = Aba_metricas['A18'].value
Valor_area_responsavel_nps_colaboradores = Aba_metricas['B18'].value
Valor_setor_responsavel_nps_colaboradores = Aba_metricas['C18'].value
Valor_frente2_nps_colaboradores = Aba_metricas['D18'].value
Valor_indicador_nps_colaboradores = Aba_metricas['E18'].value
Valor_meta_nps_colaboradores = Aba_metricas['F18'].value
Valor_ruim_nps_colaboradores = Aba_metricas['G18'].value
Valor_regular_nps_colaboradores = Aba_metricas['H18'].value
Valor_otimo_nps_colaboradores = Aba_metricas['I18'].value
Valor_peso_nps_colaboradores = Aba_metricas['J18'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_nps_colaboradores,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_nps_colaboradores,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_nps_colaboradores,
    'FRENTE 2': Valor_frente2_nps_colaboradores,
    'INDICADOR': Valor_indicador_nps_colaboradores,
    'META': Valor_meta_nps_colaboradores,
    'RUIM': Valor_ruim_nps_colaboradores,
    'REGULAR': Valor_regular_nps_colaboradores,
    'ÓTIMO': Valor_otimo_nps_colaboradores,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_nps_colaboradores,
}

# Criar o DataFrame
df_nps_colaboradores = pd.DataFrame(dados)
nps_colaboradores = []

for mes in meses:
    if dados_excel_nps_colaboradores['Data NPS'].isna().all():
        nps_colaboradores = None
    else:
        # Filtrar os dados para o mês atual
        dados_mes_nps_colaboradores = dados_excel_nps_colaboradores[
            dados_excel_nps_colaboradores['Data NPS'].dt.month == mes]

        qtde_nps_colaboradores = dados_mes_nps_colaboradores['Resultado NPS'].sum()

        if qtde_nps_colaboradores == 0:
            nps_colaboradores.append(None)
        else:
            nps_colaboradores.append(qtde_nps_colaboradores)

df_nps_colaboradores['Valor'] = nps_colaboradores

Tratamento_nps_colaboradores = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_nps_colaboradores.xlsx'
df_nps_colaboradores.to_excel(Tratamento_nps_colaboradores,index=False)

############################################# NPS PAIS E RESPONSAVEIS ##########################################################
############################# INDICADOR : Resultado da nota NPS Pais e Responsáveis

caminho_arquivo_nps_pais_responsaveis = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.5. NPS Pais e Responsáveis\23. NPS Pais e Responsáveis.xlsx'
nome_aba_nps_pais_responsaveis = 'Planilha1'
dados_excel_nps_pais_responsaveis = pd.read_excel(caminho_arquivo_nps_pais_responsaveis, sheet_name=nome_aba_nps_pais_responsaveis)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\1. Novos Indicadores Escola.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha1']

# Puxar o valor da célula A1
Valor_frente1_nps_pais_e_responsaveis = Aba_metricas['A19'].value
Valor_area_responsavel_nps_pais_e_responsaveis = Aba_metricas['B19'].value
Valor_setor_responsavel_nps_pais_e_responsaveis = Aba_metricas['C19'].value
Valor_frente2_nps_pais_e_responsaveis = Aba_metricas['D19'].value
Valor_indicador_nps_pais_e_responsaveis = Aba_metricas['E19'].value
Valor_meta_nps_pais_e_responsaveis = Aba_metricas['F19'].value
Valor_ruim_nps_pais_e_responsaveis = Aba_metricas['G19'].value
Valor_regular_nps_pais_e_responsaveis = Aba_metricas['H19'].value
Valor_otimo_nps_pais_e_responsaveis = Aba_metricas['I19'].value
Valor_peso_nps_pais_e_responsaveis = Aba_metricas['J19'].value

# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_nps_pais_e_responsaveis,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_nps_pais_e_responsaveis,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_nps_pais_e_responsaveis,
    'FRENTE 2': Valor_frente2_nps_pais_e_responsaveis,
    'INDICADOR': Valor_indicador_nps_pais_e_responsaveis,
    'META': Valor_meta_nps_pais_e_responsaveis,
    'RUIM': Valor_ruim_nps_pais_e_responsaveis,
    'REGULAR': Valor_regular_nps_pais_e_responsaveis,
    'ÓTIMO': Valor_otimo_nps_pais_e_responsaveis,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_nps_pais_e_responsaveis,
}

# Criar o DataFrame
df_nps_pais_e_responsaveis = pd.DataFrame(dados)
nps_pais_e_responsaveis = []

for mes in meses:
    if dados_excel_nps_pais_responsaveis['Data NPS'].isna().all():
        nps_pais_e_responsaveis = None
    else:
        # Filtrar os dados para o mês atual
        dados_mes_nps_pais_responsaveis = dados_excel_nps_pais_responsaveis[dados_excel_nps_pais_responsaveis['Data NPS'].dt.month == mes]

        qtde_nps_pais_responsaveis = dados_mes_nps_pais_responsaveis['Resultado NPS'].sum()

        if qtde_nps_pais_responsaveis == 0:
            nps_pais_e_responsaveis.append(None)
        else:
            nps_pais_e_responsaveis.append(qtde_nps_pais_responsaveis)

df_nps_pais_e_responsaveis['Valor'] = nps_pais_e_responsaveis

Tratamento_nps_pais_e_responsaveis = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_nps_pais_e_responsaveis.xlsx'
df_nps_pais_e_responsaveis.to_excel(Tratamento_nps_pais_e_responsaveis,index=False)

############################################# 22 - Financeiro_ebtida ##########################################################

caminho_arquivo_ebtida = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.3. Financeiro - MVM\22 - Financeiro_ebtida.xlsx'
nome_aba_ebtida = 'EBTIDA'
dados_excel_ebtida = pd.read_excel(caminho_arquivo_ebtida, sheet_name=nome_aba_ebtida)

# Criar a lista de meses de 1 a 12
meses = list(range(1, 13))
ano_atual = datetime.now().year

# Carregar o arquivo de métrica
planilha = openpyxl.load_workbook(r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.3. Métricas\2. Régua Indicadores Escola - Financeiro.xlsx')

# Selecionar a planilha desejada
Aba_metricas = planilha['Planilha2']

# Puxar o valor da célula A1
Valor_frente1_ebtida = Aba_metricas['A10'].value
Valor_area_responsavel_ebtida = Aba_metricas['B10'].value
Valor_setor_responsavel_ebtida = Aba_metricas['C10'].value
Valor_frente2_ebtida = Aba_metricas['D10'].value
Valor_indicador_ebtida = Aba_metricas['E10'].value
Valor_meta_ebtida = Aba_metricas['F10'].value
Valor_ruim_ebtida = Aba_metricas['G10'].value
Valor_regular_ebtida = Aba_metricas['H10'].value
Valor_otimo_ebtida = Aba_metricas['I10'].value
Valor_peso_ebtida = Aba_metricas['J10'].value


# Criar os dados para a tabela
dados = {
    'Ano': [ano_atual] * len(meses),
    'Mês': meses,
    'Unidade': ["MVME"] * len(meses),
    'FRENTE 1': Valor_frente1_ebtida,
    'ÁREA RESPONSÁVEL RESULTADO': Valor_area_responsavel_ebtida,
    'SETOR RESPONSÁVEL COLETA DADO': Valor_setor_responsavel_ebtida,
    'FRENTE 2': Valor_frente2_ebtida,
    'INDICADOR': Valor_indicador_ebtida,
    'META': Valor_meta_ebtida,
    'RUIM': Valor_ruim_ebtida,
    'REGULAR': Valor_regular_ebtida,
    'ÓTIMO': Valor_otimo_ebtida,
    'Valor': [0] * len(meses),
    'Tipo':["Porcentagem"] * len(meses),
    'Peso': Valor_peso_ebtida,
}

# Criar o DataFrame
df_ebtida = pd.DataFrame(dados)
financeiro_ebtida = []

for mes in meses:
    if dados_excel_ebtida['Mês'].isna().all():
        financeiro_ebtida = None
    else:
        # Filtrar os dados para o mês atual e para o custo "Pessoal Pedagógico"
        dados_ebtida = dados_excel_ebtida[(dados_excel_ebtida['Mês'] == mes)]

        # Obter o valor do custo "Pessoal Pedagógico" para o mês atual
        financeiros_ebtida = dados_ebtida['Valor'].sum()

        financeiro_ebtida.append(financeiros_ebtida)


df_ebtida['Valor'] = financeiro_ebtida

# Formatar a coluna "Valor" como moeda brasileira (R$)
df_ebtida['Valor'] = df_ebtida['Valor'].apply(lambda x: '{:.2%}'.format(x) if x is not None else None)

Tratamento_ebtida = r'G:\.shortcut-targets-by-id\1kArAZwgCxrjbQwQOPEzeJLtMUll3VVJ7\13. Dados\13.2. RMR\13.2.2. Escolas\13.2.2.2. MVM\13.2.2.2.2. Quantitativo\13.2.2.2.2.1. Base\13.2.2.2.2.1.2. Base Escolas\13.2.2.2.2.1.2.1. Excel\Tratamento_ebtida.xlsx'
df_ebtida.to_excel(Tratamento_ebtida,index=False)